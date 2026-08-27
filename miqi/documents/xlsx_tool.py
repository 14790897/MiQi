"""Excel (.xlsx) read/write tools."""

from __future__ import annotations

from pathlib import Path
from typing import Any

from miqi.agent.tools.base import Tool
from miqi.agent.tools.filesystem import _persist_tracked_file
from miqi.documents.path_utils import (
    enforce_boundary,
    ensure_suffix,
    raw_output_path,
    resolve_output_path,
)


def _cell_value_matches(actual: Any, expected: Any) -> bool:
    if actual == expected:
        return True
    try:
        return float(actual) == float(expected)
    except (TypeError, ValueError):
        return str(actual) == str(expected)


def _find_vertical_range(ws: Any, values: list[Any]) -> tuple[int, int, int, int] | None:
    if not values:
        return None
    length = len(values)
    for col in range(1, ws.max_column + 1):
        for start_row in range(1, ws.max_row - length + 2):
            if all(
                _cell_value_matches(ws.cell(start_row + offset, col).value, value)
                for offset, value in enumerate(values)
            ):
                return col, start_row, col, start_row + length - 1
    return None


def _find_named_series_range(
    ws: Any,
    name: Any,
    values: list[Any],
) -> tuple[int, int, int, int] | None:
    if not values:
        return None
    for row in range(1, ws.max_row + 1):
        for col in range(1, ws.max_column + 1):
            if not _cell_value_matches(ws.cell(row, col).value, name):
                continue
            if row + len(values) > ws.max_row:
                continue
            if all(
                _cell_value_matches(ws.cell(row + 1 + offset, col).value, value)
                for offset, value in enumerate(values)
            ):
                return col, row, col, row + len(values)
    return None


def _add_chart(ws: Any, chart_spec: dict[str, Any]) -> bool:
    from openpyxl.chart import BarChart, LineChart, PieChart, Reference

    chart_type = str(chart_spec.get("type", "bar")).lower()
    chart = (
        LineChart()
        if chart_type == "line"
        else PieChart()
        if chart_type == "pie"
        else BarChart()
    )
    chart.title = str(chart_spec.get("title", ""))
    anchor = str(chart_spec.get("anchor") or chart_spec.get("anchor_cell") or "E2")

    data_range = chart_spec.get("data_range")
    if data_range:
        data = Reference(ws, range_string=f"'{ws.title}'!{data_range}")
        chart.add_data(
            data,
            titles_from_data=bool(chart_spec.get("titles_from_data", True)),
        )
    else:
        series_specs = chart_spec.get("series") or []
        if not isinstance(series_specs, list) or not series_specs:
            return False
        added_series = False
        for series in series_specs:
            if not isinstance(series, dict):
                continue
            found = _find_named_series_range(
                ws,
                series.get("name"),
                list(series.get("values") or []),
            )
            if found is None:
                continue
            min_col, min_row, max_col, max_row = found
            data = Reference(
                ws,
                min_col=min_col,
                min_row=min_row,
                max_col=max_col,
                max_row=max_row,
            )
            chart.add_data(data, titles_from_data=True)
            added_series = True
        if not added_series:
            return False

    category_range = chart_spec.get("category_range")
    if category_range:
        cats = Reference(ws, range_string=f"'{ws.title}'!{category_range}")
        chart.set_categories(cats)
    elif isinstance(chart_spec.get("categories"), list):
        found = _find_vertical_range(ws, list(chart_spec["categories"]))
        if found is not None:
            min_col, min_row, max_col, max_row = found
            cats = Reference(
                ws,
                min_col=min_col,
                min_row=min_row,
                max_col=max_col,
                max_row=max_row,
            )
            chart.set_categories(cats)

    ws.add_chart(chart, anchor)
    return True


class XlsxReadTool(Tool):
    """Read data from an Excel (.xlsx) spreadsheet."""

    name = "xlsx_read"
    description = "Read and extract data from an Excel (.xlsx) file."

    def __init__(
        self,
        workspace: Path | None = None,
        allowed_dir: Path | None = None,
    ):
        self._workspace = workspace
        self._allowed_dir = allowed_dir

    @property
    def parameters(self) -> dict[str, Any]:
        return {
            "type": "object",
            "properties": {
                "file_path": {
                    "type": "string",
                    "description": "Path to the .xlsx file to read. Alias for filename.",
                },
                "filename": {
                    "type": "string",
                    "description": "Filename or relative path to the .xlsx file.",
                },
                "path": {
                    "type": "string",
                    "description": "Path to the .xlsx file to read. Alias for filename.",
                },
                "sheet_name": {
                    "type": "string",
                    "description": "Optional sheet name to read. Reads first sheet if omitted.",
                },
            },
            "anyOf": [
                {"required": ["filename"]},
                {"required": ["file_path"]},
                {"required": ["path"]},
            ],
        }

    async def execute(self, **kwargs: Any) -> str:
        _sess_key = kwargs.pop("_session_key", None)
        raw_path = raw_output_path(kwargs)
        if not raw_path.strip():
            return "Error: 必须提供 filename"
        try:
            file_path = resolve_output_path(
                raw_path, self._workspace, self._allowed_dir,
            )
            file_path = ensure_suffix(file_path, ".xlsx")
            enforce_boundary(file_path, self._allowed_dir, self._workspace)
        except PermissionError as e:
            return f"Error: 权限被拒绝：{e}"
        except ValueError as e:
            return f"Error: {e}"
        if not file_path.exists():
            return f"Error: 文件不存在：{file_path}"
        try:
            from openpyxl import load_workbook

            wb = load_workbook(str(file_path), read_only=True, data_only=True)
            sheet_name = kwargs.get("sheet_name")
            if sheet_name and sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
            else:
                ws = wb.active

            lines = [f"Sheet: {ws.title} ({ws.max_row or '?'} rows x {ws.max_column or '?'} cols)", ""]
            max_row = ws.max_row or 2000
            for row in ws.iter_rows(values_only=True, max_row=min(max_row, 200)):
                cells = [str(c) if c is not None else "" for c in row]
                lines.append(" | ".join(cells))

            if max_row > 200:
                lines.append(f"\n... ({max_row - 200} more rows)")

            wb.close()
            return "\n".join(lines)
        except Exception as e:
            return f"Error reading {file_path.name}: {e}"


class CreateXlsxTool(Tool):
    """Create an Excel (.xlsx) spreadsheet."""

    name = "create_xlsx"
    description = (
        "Create an Excel (.xlsx) workbook in the workspace files directory. "
        "Supports multiple sheets, formulas, and simple charts."
    )

    def __init__(
        self,
        workspace: Path | None = None,
        allowed_dir: Path | None = None,
    ):
        self._workspace = workspace
        self._allowed_dir = allowed_dir

    @property
    def parameters(self) -> dict[str, Any]:
        return {
            "type": "object",
            "properties": {
                "file_path": {
                    "type": "string",
                    "description": "Path for the output .xlsx file. Alias for filename.",
                },
                "filename": {
                    "type": "string",
                    "description": "Filename or relative path for the output .xlsx file.",
                },
                "path": {
                    "type": "string",
                    "description": "Path for the output .xlsx file. Alias for filename.",
                },
                "sheets": {
                    "description": (
                        "Either {sheet_name: rows} or "
                        "[{name: string, rows: [[...]], charts: [...]}]. "
                        "Formula strings beginning with '=' are preserved."
                    ),
                },
                "rows": {
                    "type": "array",
                    "items": {"type": "array"},
                    "description": (
                        "Rows for a single-sheet workbook. Use this when the "
                        "user asks for one sheet and no sheets array is needed."
                    ),
                },
                "sheet_name": {
                    "type": "string",
                    "description": "Sheet name to use with top-level rows.",
                },
                "charts": {
                    "type": "array",
                    "items": {"type": "object"},
                    "description": "Charts to add when using top-level rows.",
                },
            },
            "anyOf": [
                {"required": ["filename"]},
                {"required": ["file_path"]},
                {"required": ["path"]},
            ],
        }

    async def execute(self, **kwargs: Any) -> str:
        from openpyxl import Workbook

        _sess_key = kwargs.pop("_session_key", None)
        raw_path = raw_output_path(kwargs)
        sheets = kwargs.get("sheets")
        rows = kwargs.get("rows")
        if not raw_path.strip():
            return "Error: 必须提供 filename"

        try:
            file_path = resolve_output_path(
                raw_path, self._workspace, self._allowed_dir,
            )
            file_path = ensure_suffix(file_path, ".xlsx")
            enforce_boundary(file_path, self._allowed_dir, self._workspace)
        except PermissionError as e:
            return f"Error: 权限被拒绝：{e}"
        except ValueError as e:
            return f"Error: {e}"
        if sheets is None and rows is None:
            return "Error: 必须提供 sheets 或 rows"

        try:
            wb = Workbook()
            wb.remove(wb.active)

            if sheets is None:
                sheet_specs = [
                    {
                        "name": kwargs.get("sheet_name") or kwargs.get("name") or "Sheet1",
                        "rows": rows,
                        "charts": kwargs.get("charts") or [],
                    }
                ]
            elif isinstance(sheets, dict):
                sheet_specs = [
                    {"name": sheet_name, "rows": rows}
                    for sheet_name, rows in sheets.items()
                ]
            elif isinstance(sheets, list):
                sheet_specs = sheets
            else:
                return "Error: sheets 必须是对象或数组"

            if not sheet_specs:
                return "Error: sheets 至少需要包含一个工作表"

            created_sheets = 0
            for index, spec in enumerate(sheet_specs):
                if not isinstance(spec, dict):
                    continue
                sheet_name = str(spec.get("name") or f"Sheet{index + 1}")[:31]
                ws = wb.create_sheet(sheet_name, index)
                created_sheets += 1
                for row_data in spec.get("rows", []) or []:
                    ws.append(row_data)
                for chart_spec in spec.get("charts", []) or []:
                    if not isinstance(chart_spec, dict):
                        continue
                    _add_chart(ws, chart_spec)

            if created_sheets == 0:
                return "Error: sheets 必须包含工作表对象"

            file_path.parent.mkdir(parents=True, exist_ok=True)
            wb.save(str(file_path))
            _persist_tracked_file(self._workspace, file_path, op="write", session_key=_sess_key)
            return f"Created: {file_path} ({len(wb.sheetnames)} sheet(s))"
        except Exception as e:
            return f"Error writing {raw_path}: {e}"


class XlsxWriteTool(CreateXlsxTool):
    """Backward-compatible alias for create_xlsx."""

    name = "xlsx_write"
    description = "Create a new Excel (.xlsx) file. Prefer create_xlsx for new calls."


class AppendXlsxTool(Tool):
    """Append rows to an existing Excel (.xlsx) workbook."""

    name = "append_xlsx"
    description = (
        "Append rows to an existing Excel (.xlsx) workbook in the workspace files directory."
    )

    def __init__(
        self,
        workspace: Path | None = None,
        allowed_dir: Path | None = None,
    ):
        self._workspace = workspace
        self._allowed_dir = allowed_dir

    @property
    def parameters(self) -> dict[str, Any]:
        return {
            "type": "object",
            "properties": {
                "file_path": {
                    "type": "string",
                    "description": "Path to the .xlsx file. Alias for filename.",
                },
                "filename": {
                    "type": "string",
                    "description": "Filename or relative path to the .xlsx file.",
                },
                "path": {
                    "type": "string",
                    "description": "Path to the .xlsx file. Alias for filename.",
                },
                "sheet_name": {
                    "type": "string",
                    "description": "Sheet to append to. Defaults to the active sheet.",
                },
                "rows": {
                    "type": "array",
                    "items": {"type": "array"},
                    "description": "Rows to append.",
                },
                "create_sheet": {
                    "type": "boolean",
                    "description": "Create sheet_name if it does not exist.",
                },
            },
            "required": ["rows"],
            "anyOf": [
                {"required": ["filename"]},
                {"required": ["file_path"]},
                {"required": ["path"]},
            ],
        }

    async def execute(self, **kwargs: Any) -> str:
        from openpyxl import load_workbook

        _sess_key = kwargs.pop("_session_key", None)
        raw_path = raw_output_path(kwargs)
        if not raw_path.strip():
            return "Error: 必须提供 filename"

        try:
            file_path = resolve_output_path(
                raw_path, self._workspace, self._allowed_dir,
            )
            file_path = ensure_suffix(file_path, ".xlsx")
            enforce_boundary(file_path, self._allowed_dir, self._workspace)
        except PermissionError as e:
            return f"Error: 权限被拒绝：{e}"
        except ValueError as e:
            return f"Error: {e}"

        if not file_path.exists():
            return f"Error: 文件不存在：{file_path}"

        rows = kwargs.get("rows") or []
        if not rows:
            return "Error: 必须提供 rows"

        try:
            wb = load_workbook(str(file_path))
            sheet_name = kwargs.get("sheet_name")
            if sheet_name:
                if sheet_name in wb.sheetnames:
                    ws = wb[sheet_name]
                elif kwargs.get("create_sheet", False):
                    ws = wb.create_sheet(str(sheet_name)[:31])
                else:
                    wb.close()
                    return f"Error: 未找到工作表：{sheet_name}"
            else:
                ws = wb.active

            for row in rows:
                ws.append(row)

            wb.save(str(file_path))
            wb.close()
            _persist_tracked_file(self._workspace, file_path, op="edit", session_key=_sess_key)
            return f"Appended: {file_path} ({len(rows)} row(s) to {ws.title})"
        except Exception as e:
            return f"Error editing {raw_path}: {e}"
